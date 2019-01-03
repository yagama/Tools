from openpyxl import Workbook
from openpyxl import load_workbook

# 讀檔
wb = load_workbook('bom.xlsx')
# 開新檔案
newwb = Workbook()

# 指定工作表
sheet = wb.get_sheet_by_name('bom')
new = newwb.active

#先定義Sheet1要使用的A, B欄row index初始值
global_A_col = 1	
global_B_col = 1
# 印出bom J列的所有內容
for i in sheet["J"]:
	rawstr = str(i.value)
	#計算該欄位以 , 區分的陣列長度
	str_len = len(rawstr.split(",")) 
	for j in range(0,str_len):
		#寫入Sheet1的A1
		new['A'+str(global_B_col)] = sheet['D'+str(global_A_col)].value
		#寫入Sheet1的B1
		new['B'+str(global_B_col)] = rawstr.split(",")[j]
		global_B_col = global_B_col + 1
	#Bom J列第一行處理完後，再把整個Index +1
	global_A_col = global_A_col + 1
		
#另存新檔，檔名為Final.xlsx
newwb.save('Output.xlsx')
